import openpyxl

# Crear un nuevo libro de trabajo de Excel
workbook = openpyxl.Workbook()

# Crear una nueva hoja llamada "Ventas"
ventas_sheet = workbook.active
ventas_sheet.title = "Ventas"

# Crear encabezados de columna
ventas_sheet["A1"] = "Categoría"
ventas_sheet["B1"] = "Sabor"
ventas_sheet["C1"] = "Cantidad"
ventas_sheet["D1"] = "Precio"

# Definir precios de cada categoría
precios = {"Mini": 3, "Holanda": 5}

while True:
    # Pedir al usuario que seleccione la categoría para agregar ventas
    categoria = input("Ingrese la categoría (Mini/Holanda): ")

    # Verificar que la categoría ingresada sea válida
    while categoria not in ["Mini", "Holanda"]:
        categoria = input("Categoría no válida. Ingrese Mini o Holanda: ")

    # Pedir al usuario que ingrese el sabor y la cantidad de paletas vendidas
    while True:
        sabor = input("Ingrese el sabor de la paleta (Enter para terminar): ")
        if sabor == "":
            break
        cantidad = int(input("Ingrese la cantidad de paletas vendidas: "))
        precio = precios[categoria]
        ventas_sheet.append([categoria, sabor, cantidad, precio])

        # Preguntar si se quiere seguir agregando ventas en la misma categoría
        seguir_agregando = input(f"Desea seguir agregando ventas en la categoría {categoria}? (S/N): ")
        if seguir_agregando.upper() == "N":
            break

    # Preguntar si se quiere agregar más ventas de otra categoría
    seguir_agregando_categoria = input("Desea agregar ventas en otra categoría? (S/N): ")
    if seguir_agregando_categoria.upper() == "N":
        break

# Calcular el total de ventas y cantidad de paletas vendidas
total_ventas = 0
total_cantidad = 0
for row in ventas_sheet.iter_rows(min_row=2, values_only=True):
    total_ventas += row[2] * row[3]
    total_cantidad += row[2]

# Agregar una hoja de resumen con los totales de ventas y cantidad
resumen_sheet = workbook.create_sheet("Resumen")
resumen_sheet["A1"] = "Fecha"
resumen_sheet["B1"] = "Total Paletas Mini"
resumen_sheet["C1"] = "Total Paletas Holanda"
resumen_sheet["D1"] = "Total Ventas"
resumen_sheet.append([input("Ingrese la fecha: "), "=SUMIF(Ventas!A:A,\"Mini\",Ventas!C:C)", "=SUMIF(Ventas!A:A,\"Holanda\",Ventas!C:C)", total_ventas])

# Guardar el archivo de Excel
filename = input("Ingrese el nombre del archivo: ")
workbook.save(f"{filename}.xlsx")

print("Archivo guardado correctamente.")
